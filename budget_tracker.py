import os
import json
from datetime import date
import openpyxl
from groq import Groq

from dotenv import load_dotenv

load_dotenv()

client = Groq(api_key=os.environ.get("GROQ_API_KEY"))
EXCEL_FILE = "budget.xlsx"


def get_or_create_workbook():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Budget"
        ws.append(["Date", "Type", "Category", "Amount", "Description"])
        wb.save(EXCEL_FILE)
    return wb


def entry_agent(user_input: str) -> dict:
    """Agent 1: Parses user input and extracts structured budget entries."""
    response = client.chat.completions.create(
        model="llama-3.1-8b-instant",
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "system",
                "content": (
                    "Extract budget entries from user input. "
                    'Return ONLY a JSON object with a single key \'entries\' containing an array of objects with keys: '
                    "type (income/expense), category, amount (number), description."
                ),
            },
            {"role": "user", "content": user_input},
        ],
    )
    raw = response.choices[0].message.content.strip()
    return json.loads(raw)["entries"]


def save_entries(entries: list):
    wb = get_or_create_workbook()
    ws = wb["Budget"]
    for e in entries:
        ws.append([str(date.today()), e["type"], e["category"], e["amount"], e["description"]])
    wb.save(EXCEL_FILE)
    print(f"✅ Saved {len(entries)} entry/entries to {EXCEL_FILE}")


def load_summary() -> dict:
    wb = get_or_create_workbook()
    ws = wb["Budget"]
    summary = {"income": 0, "expenses": {}}
    for row in ws.iter_rows(min_row=2, values_only=True):
        _, typ, category, amount, _ = row
        if not amount:
            continue
        if typ == "income":
            summary["income"] += amount
        elif typ == "expense":
            summary["expenses"][category] = summary["expenses"].get(category, 0) + amount
    return summary


def advisor_agent(summary: dict) -> str:
    """Agent 2: Analyzes spending and suggests savings tips."""
    response = client.chat.completions.create(
        model="llama3-8b-8192",
        messages=[
            {
                "role": "system",
                "content": (
                    "You are a personal finance advisor. "
                    "Analyze the budget summary and give concise, actionable savings tips. "
                    "Mention percentage breakdowns where relevant. Be direct and brief."
                ),
            },
            {"role": "user", "content": f"Budget summary: {json.dumps(summary)}"},
        ],
    )
    return response.choices[0].message.content.strip()


def main():
    print("💰 Budget Tracker — type 'analyze' to get advice, 'quit' to exit\n")
    while True:
        user_input = input("You: ").strip()
        if not user_input:
            continue
        if user_input.lower() == "quit":
            break
        if user_input.lower() == "analyze":
            summary = load_summary()
            if not summary["expenses"] and summary["income"] == 0:
                print("⚠️  No data yet. Add some income/expenses first.\n")
                continue
            print("\n📊 Advisor Agent analyzing your budget...\n")
            advice = advisor_agent(summary)
            print(f"Advisor: {advice}\n")
        else:
            print("🔍 Entry Agent processing...\n")
            entries = entry_agent(user_input)
            save_entries(entries)
            for e in entries:
                print(f"  → [{e['type'].upper()}] {e['category']}: ${e['amount']} — {e['description']}")
            print()


if __name__ == "__main__":
    main()
