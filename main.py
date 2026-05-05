import os
import json
from datetime import date
import openpyxl
from groq import Groq
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel

load_dotenv()

app = FastAPI()
client = Groq(api_key=os.environ.get("GROQ_API_KEY"))
EXCEL_FILE = "budget.xlsx"


def get_or_create_workbook():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Budget"
        ws.append(["Date", "Type", "Category", "Amount", "Description"])
        wb.save(EXCEL_FILE)
    return wb


def load_summary() -> dict:
    wb = get_or_create_workbook()
    ws = wb["Budget"]
    summary = {"income": 0, "expenses": {}}
    for row in ws.iter_rows(min_row=2, values_only=True):
        _, typ, category, amount, _ = row
        if not amount:
            continue
        if typ == "income":
            summary["income"] += amount
        elif typ == "expense":
            summary["expenses"][category] = summary["expenses"].get(category, 0) + amount
    return summary


class UserInput(BaseModel):
    text: str


@app.get("/")
def root():
    return FileResponse("static/index.html")


@app.post("/add")
def add_entries(body: UserInput):
    response = client.chat.completions.create(
        model="llama-3.1-8b-instant",
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "system",
                "content": (
                    "Extract budget entries from user input. "
                    "Return ONLY a JSON object with a single key 'entries' containing an array of objects with keys: "
                    "type (income/expense), category, amount (number), description."
                ),
            },
            {"role": "user", "content": body.text},
        ],
    )
    entries = json.loads(response.choices[0].message.content)["entries"]
    wb = get_or_create_workbook()
    ws = wb["Budget"]
    for e in entries:
        ws.append([str(date.today()), e["type"], e["category"], e["amount"], e["description"]])
    wb.save(EXCEL_FILE)
    return {"entries": entries}


@app.get("/analyze")
def analyze():
    summary = load_summary()
    if not summary["expenses"] and summary["income"] == 0:
        raise HTTPException(status_code=400, detail="No data yet. Add some income/expenses first.")
    response = client.chat.completions.create(
        model="llama-3.1-8b-instant",
        messages=[
            {
                "role": "system",
                "content": (
                    "You are a personal finance advisor. "
                    "Analyze the budget summary and give concise, actionable savings tips. "
                    "Mention percentage breakdowns where relevant. Be direct and brief."
                ),
            },
            {"role": "user", "content": f"Budget summary: {json.dumps(summary)}"},
        ],
    )
    return {"advice": response.choices[0].message.content.strip(), "summary": summary}


@app.get("/entries")
def get_entries():
    wb = get_or_create_workbook()
    ws = wb["Budget"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append({"date": row[0], "type": row[1], "category": row[2], "amount": row[3], "description": row[4]})
    return {"entries": rows}


app.mount("/static", StaticFiles(directory="static"), name="static")
